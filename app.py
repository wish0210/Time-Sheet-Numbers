from flask import Flask, request, redirect, send_file
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
import os

app = Flask(__name__)

@app.route('/')
def upload_file():
    return '''
    <!doctype html>
    <title>Upload CSV File</title>
    <h1>Upload CSV File</h1>
    <form action="/process" method="post" enctype="multipart/form-data">
      <input type="file" name="file">
      <input type="submit" value="Upload">
    </form>
    '''

@app.route('/process', methods=['POST'])
def process_file():
    if 'file' not in request.files:
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)
    if file:
        try:
            # Read the first three lines separately
            file.seek(0)
            first_three_lines = [file.readline().decode('ISO-8859-1').strip(),
                                 file.readline().decode('ISO-8859-1').strip(),
                                 file.readline().decode('ISO-8859-1').strip()]
            file.seek(0)
            df = pd.read_csv(file, encoding='ISO-8859-1', skiprows=2, on_bad_lines='skip', header=0)
        except pd.errors.ParserError as e:
            return f"Error parsing CSV file: {e}"
        except UnicodeDecodeError as e:
            return f"Error decoding CSV file: {e}"

        try:
            output = process_data(df, first_three_lines)
        except ValueError as e:
            return f"{e}"
        
        original_filename = os.path.splitext(file.filename)[0] + "_processed.xlsx"
        return send_file(output, download_name=original_filename, as_attachment=True)

def process_data(df, first_three_lines):
    required_columns = ['EMP L NAME', 'EMP F NAME', 'DATE', 'IN', 'OUT', 'TOTAL']
    df.columns = df.columns.str.strip().str.upper()  # Strip any whitespace and convert column names to uppercase

    # Debugging: print the columns to see what is being read
    print("Columns in CSV:", df.columns.tolist())

    # Attempt to match required columns
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError("CSV file is missing required columns: " + ", ".join(missing_columns))

    df['DATE'] = pd.to_datetime(df['DATE'], errors='coerce')
    if df['DATE'].isna().any():
        raise ValueError("Invalid date format in 'DATE' column.")

    output = BytesIO()
    wb = Workbook()
    ws = wb.active

    # Split the header line into separate cells starting from column B
    header_values = first_three_lines[2].split(',')
    for col_idx, header in enumerate(header_values, start=2):
        ws.cell(row=3, column=col_idx, value=header.strip())

    # Write the first two lines to start from column B, making sure to write each value separately
    first_line_values = first_three_lines[0].split(',')
    second_line_values = first_three_lines[1].split(',')
    for col_idx, value in enumerate(first_line_values, start=2):
        ws.cell(row=1, column=col_idx, value=value.strip())
    for col_idx, value in enumerate(second_line_values, start=2):
        ws.cell(row=2, column=col_idx, value=value.strip())

    row = 4  # Start writing data from the fourth row
    employee_counter = 1

    # Group by employee but maintain the order
    for (lname, fname), group in df.groupby(['EMP L NAME', 'EMP F NAME'], sort=False):
        ws.cell(row=row, column=1, value=employee_counter)
        ws.merge_cells(start_row=row, start_column=1, end_row=row + len(group) - 1, end_column=1)
        employee_counter += 1
        
        # Determine the first Monday on or before the earliest date in the group
        first_date = group['DATE'].min()
        first_monday = first_date - pd.DateOffset(days=(first_date.weekday() - 0) % 7)
        current_week = pd.date_range(start=first_monday, periods=7, freq='D')

        for date in current_week:
            entries_for_date = group[group['DATE'] == date]
            if not entries_for_date.empty:
                for _, entry in entries_for_date.iterrows():
                    ws.cell(row=row, column=2, value=entry['EMP L NAME'])
                    ws.cell(row=row, column=3, value=entry['EMP F NAME'])
                    ws.cell(row=row, column=4, value=entry['DATE'].strftime('%Y-%m-%d'))
                    ws.cell(row=row, column=5, value=entry['IN'])
                    ws.cell(row=row, column=6, value=entry['OUT'])
                    ws.cell(row=row, column=7, value=entry['TOTAL'])
                    row += 1
            else:
                ws.cell(row=row, column=2, value=lname)
                ws.cell(row=row, column=3, value=fname)
                ws.cell(row=row, column=4, value=date.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=5, value="Off")
                ws.cell(row=row, column=6, value="Off")
                ws.cell(row=row, column=7, value="")
                row += 1
        row += 1

    wb.save(output)
    output.seek(0)
    return output

if __name__ == '__main__':
    app.run(debug=True)
